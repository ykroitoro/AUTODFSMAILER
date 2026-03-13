import os
import io
import copy
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
from dotenv import load_dotenv
import dropbox
from dropbox.exceptions import ApiError

# ─────────────────────────────────────────────
#  CONFIG
# ─────────────────────────────────────────────
load_dotenv()

app_key      = os.getenv("DROPBOX_APP_KEY")
app_secret   = os.getenv("DROPBOX_APP_SECRET")
refresh_token = os.getenv("DROPBOX_REFRESH_TOKEN")

# Hardcoded — always calculate price, always delete Cost/Product ID
ans1 = "yes"   # "yes" = calculate Price column
ans2 = "yes"   # "yes" = delete Cost and Product ID columns

# Dropbox paths
INPUT_PATH         = "/AUTODFSMAILER/DFS_LIST.XLSX"
MENU_TEMPLATE_PATH = "/AUTODFSMAILER/DELL_LIST_MENU.xlsx"

timestamp       = datetime.now().strftime("%Y%m%d_%H%M%S")
OUTPUT_FILENAME = f"DELL_LIST_GRADE_A_{timestamp}.xlsx"
OUTPUT_PATH     = f"/AUTODFSMAILER/{OUTPUT_FILENAME}"

# Local temp paths (Railway ephemeral disk — only used mid-run)
LOCAL_INPUT      = f"/tmp/DFS_LIST.xlsx"
LOCAL_MENU       = f"/tmp/DELL_LIST_MENU.xlsx"
LOCAL_OUTPUT     = f"/tmp/{OUTPUT_FILENAME}"


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def sanitize_sheet_title(title):
    if not isinstance(title, str):
        title = str(title) if title is not None else "Unnamed"
    title = title.strip()[:31]
    for c in ['\\', '/', '*', '[', ']', ':', '?']:
        title = title.replace(c, '')
    return title or "Unnamed"


# ─────────────────────────────────────────────
#  STEP 1 — Connect to Dropbox
# ─────────────────────────────────────────────
print("Connecting to Dropbox...")
dbx = dropbox.Dropbox(
    app_key=app_key,
    app_secret=app_secret,
    oauth2_refresh_token=refresh_token
)
dbx.users_get_current_account()   # validates credentials — raises AuthError if wrong
print("Dropbox connected OK")


# ─────────────────────────────────────────────
#  STEP 2 — Download input files from Dropbox
# ─────────────────────────────────────────────
print(f"Downloading input file: {INPUT_PATH}")
_, res = dbx.files_download(INPUT_PATH)
with open(LOCAL_INPUT, "wb") as f:
    f.write(res.content)
print("Input file downloaded OK")

print(f"Downloading menu template: {MENU_TEMPLATE_PATH}")
_, res = dbx.files_download(MENU_TEMPLATE_PATH)
with open(LOCAL_MENU, "wb") as f:
    f.write(res.content)
print("Menu template downloaded OK")


# ─────────────────────────────────────────────
#  STEP 3 — Load & clean the data
# ─────────────────────────────────────────────
print("Loading data...")
df = pd.read_excel(LOCAL_INPUT, sheet_name='Worksheet')
df.fillna('', inplace=True)

# Remove test rows
remove_values = ['TEST30', 'chargeTestFinale']
found_to_remove = df['Model'].isin(remove_values).any()
if found_to_remove:
    df = df[~df['Model'].isin(remove_values)]
    print("Removed test rows: TEST30 / chargeTestFinale")
else:
    print("No test rows found")

# Fix typo in Model column
df['Model'] = df['Model'].astype(str).str.replace('Suface', 'Surface', regex=False)

# Create BRAND column
df['BRAND'] = df['Model'].apply(
    lambda x: next((b for b in ['Dell', 'HP', 'Lenovo', 'Microsoft'] if isinstance(x, str) and b in x), '')
)
df.insert(df.columns.get_loc('Product ID') + 1, 'BRAND', df.pop('BRAND'))

# Create Model Name column
df['Model Name'] = df['Model'].apply(
    lambda x: str(x).split()[1] if isinstance(x, str) and len(str(x).split()) > 1 else ''
)
df.insert(df.columns.get_loc('Model') + 1, 'Model Name', df.pop('Model Name'))

# Create Model Number column
df['Model Number'] = df['Model'].apply(
    lambda x: str(x).split()[2] if isinstance(x, str) and len(str(x).split()) > 2 else ''
)
df.insert(df.columns.get_loc('Model Name') + 1, 'Model Number', df.pop('Model Number'))

# Create Form Factor column (UFF included)
df['Form Factor'] = df['Description'].apply(
    lambda x: next((f for f in ['SFF', 'MFF', 'MT', 'Tower', 'AIO', 'NB', 'UFF'] if isinstance(x, str) and f in x), 'NB')
)
df.insert(df.columns.get_loc('Model Number') + 1, 'Form Factor', df.pop('Form Factor'))

# Create CPU column
df['CPU'] = df['Processor'].apply(
    lambda x: x[14:26] if isinstance(x, str) and 'Ultra' in x
    else x[9:28]  if isinstance(x, str) and 'Pentium' in x
    else x[9:24]  if isinstance(x, str) and 'Xeon' in x
    else x[14:22] if isinstance(x, str) and 'Intel' in x
    else x[7:20]  if isinstance(x, str) and 'AMD' in x
    else ''
)

# Trim RAM to 2 chars (e.g. "16" from "16GB")
df['RAM'] = df['RAM'].apply(lambda x: x[:2] if isinstance(x, str) else x)

# Create ODD column
df['ODD'] = df['Description'].apply(
    lambda x: 'DVDRW' if isinstance(x, str) and 'DVDRW' in x
    else ('DVD' if isinstance(x, str) and 'DVD' in x else '')
)
df.insert(df.columns.get_loc('Form Factor') + 1, 'ODD', df.pop('ODD'))

# Create HDD column
df['HDD'] = df['Description'].apply(
    lambda x: next((s for s in ['128GB', '256GB', '500GB', '512GB', '250GB', '320GB', '1000GB', '1024GB', '768GB'] if isinstance(x, str) and s in x), '')
)
df.insert(df.columns.get_loc('ODD') + 1, 'HDD', df.pop('HDD'))

# Create Screen Size column
df['Screen_Size'] = df['Description'].apply(
    lambda x: next((s for s in ['14-in', '13.3-in', '13.5-in', '12.3-in', '15.6-in', '15,2', '13,3', '10-in', '12.5-in', '16-in', '11.6-in', '2N1', '17.3', '19.5-in', 'TAB', '24-in'] if isinstance(x, str) and s in x), '')
)
df.insert(df.columns.get_loc('HDD') + 1, 'Screen_Size', df.pop('Screen_Size'))

# Create Screen Type column
df['Screen_Type'] = df['Description'].apply(
    lambda x: next((s for s in ['NoTCH', 'TCH'] if isinstance(x, str) and s in x), '')
)
df.insert(df.columns.get_loc('Screen_Size') + 1, 'Screen_Type', df.pop('Screen_Type'))

# Create Camera column
df['Camera'] = df['Description'].apply(
    lambda x: next((s for s in ['NoCAM', 'CAM'] if isinstance(x, str) and s in x), '')
)
df.insert(df.columns.get_loc('Screen_Type') + 1, 'Camera', df.pop('Camera'))

# Move CPU and RAM right after Model Number
df.insert(df.columns.get_loc('Model Number') + 1, 'CPU', df.pop('CPU'))
df.insert(df.columns.get_loc('Model Number') + 2, 'RAM', df.pop('RAM'))

# Drop raw Processor column
df.drop(columns=['Processor'], inplace=True)

# Sort
df.sort_values(by=['Model Name', 'Model Number', 'Form Factor', 'CPU', 'RAM', 'HDD'], inplace=True)

total_lines = len(df)
print(f"Data processed: {total_lines} rows")


# ─────────────────────────────────────────────
#  STEP 4 — Save to local temp Excel
# ─────────────────────────────────────────────
df.to_excel(LOCAL_OUTPUT, index=False)

wb = load_workbook(LOCAL_OUTPUT)
ws = wb.active

# Adjust column widths
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

# Rename header cells
ws.cell(row=1, column=5,  value='Model Number')
ws.cell(row=1, column=15, value='Cost')

# Insert Price column after Cost (col 15)
col_idx_cost = 15
ws.insert_cols(col_idx_cost + 1)
ws.cell(row=1, column=col_idx_cost + 1, value='Price' if ans1 == 'yes' else 'Bid Price')

if ans1 == 'yes':
    for row in range(2, ws.max_row + 1):
        cost = ws.cell(row=row, column=col_idx_cost).value
        if isinstance(cost, (int, float)):
            price = round(((cost / 0.75) + 10) / 5, 0) * 5
            price_cell = ws.cell(row=row, column=col_idx_cost + 1, value=price)
            price_cell.number_format = '$#,##0.00'

# OS column → Windows 10 Professional if blank
for row in range(2, ws.max_row + 1):
    os_val = ws.cell(row=row, column=17).value
    if os_val == 'No Operating System':
        ws.cell(row=row, column=17).value = 'Windows 10 Professional'

# Screen Size default
for row in range(2, ws.max_row + 1):
    if not ws.cell(row=row, column=11).value:
        ws.cell(row=row, column=11).value = 'Desktop Only'

# Screen Type default
for row in range(2, ws.max_row + 1):
    if not ws.cell(row=row, column=12).value:
        ws.cell(row=row, column=12).value = 'Not Specified'

# Delete Cost and Product ID columns if ans2 == 'yes'
if ans2 == 'yes':
    ws.delete_cols(col_idx_cost)       # Cost
    ws.delete_cols(1)                  # Product ID
    print("Cost and Product ID columns deleted")
else:
    print("Cost and Product ID columns retained")

# Bold/center J1
ws['J1'].font = Font(bold=True)
ws['J1'].alignment = Alignment(horizontal='center')

wb.save(LOCAL_OUTPUT)
print("Intermediate workbook saved")


# ─────────────────────────────────────────────
#  STEP 5 — Create per-model tabs
# ─────────────────────────────────────────────
df2 = pd.read_excel(LOCAL_OUTPUT, sheet_name='Sheet1')
# Drop rows where Model Name is empty/NaN
df2 = df2[df2['Model Name'].notna() & (df2['Model Name'].astype(str).str.strip() != '')]

header = list(df2.columns)

for model in df2['Model Name'].unique():
    raw_title = sanitize_sheet_title(str(model))
    model_df = df2[df2['Model Name'] == model]
    model_ws = wb.create_sheet(title=raw_title)

    # Write header
    for col_idx, col_name in enumerate(header, 1):
        cell = model_ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

    # Write data
    for r_idx, row in enumerate(model_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            model_ws.cell(row=r_idx, column=c_idx, value=value)

    # Auto column widths
    for col in model_ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        if max_length > 0:
            model_ws.column_dimensions[col[0].column_letter].width = max_length + 5

    # Currency format for Price column
    price_col = 13 if ans2 == 'yes' else 15
    cost_col  = None if ans2 == 'yes' else 14
    for row in range(2, model_ws.max_row + 1):
        model_ws.cell(row=row, column=price_col).number_format = '$#,##0.00'
        model_ws.column_dimensions[get_column_letter(price_col)].width = 10
        if cost_col:
            model_ws.cell(row=row, column=cost_col).number_format = '$#,##0.00'
            model_ws.column_dimensions[get_column_letter(cost_col)].width = 10

# Set fixed column widths on last model sheet
if ans2 == 'yes':
    model_ws.column_dimensions['L'].width = 15
    model_ws.column_dimensions['M'].width = 30
    model_ws.column_dimensions['N'].width = 120
else:
    model_ws.column_dimensions['L'].width = 15
    model_ws.column_dimensions['M'].width = 15
    model_ws.column_dimensions['N'].width = 30
    model_ws.column_dimensions['O'].width = 120

# Remove Sheet1 and Monitor tab
for remove_name in ['Sheet1', 'Monitor']:
    if remove_name in wb.sheetnames:
        wb.remove(wb[remove_name])

wb.save(LOCAL_OUTPUT)
print(f"Per-model tabs created: {[s for s in wb.sheetnames]}")


# ─────────────────────────────────────────────
#  STEP 6 — Merge into Menu workbook
# ─────────────────────────────────────────────
print("Merging into menu workbook...")
final_wb = load_workbook(LOCAL_OUTPUT)
menu_wb  = load_workbook(LOCAL_MENU)

for sheet_name in final_wb.sheetnames:
    source_sheet = final_wb[sheet_name]
    if sheet_name in menu_wb.sheetnames:
        del menu_wb[sheet_name]
    target_sheet = menu_wb.create_sheet(sheet_name)

    # Copy cells with styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value
            if cell.has_style:
                new_cell.font       = copy.copy(cell.font)
                new_cell.fill       = copy.copy(cell.fill)
                new_cell.alignment  = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy.copy(cell.protection)
                new_cell.border     = copy.copy(cell.border)

    # Copy dimensions
    for col in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]
    for row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row] = source_sheet.row_dimensions[row]

    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    # Auto filter
    max_col = target_sheet.max_column
    target_sheet.auto_filter.ref = f"A1:{target_sheet.cell(row=1, column=max_col).coordinate}"

print("Sheets merged OK")


# ─────────────────────────────────────────────
#  STEP 7 — Build consolidated summary table on MENU sheet
# ─────────────────────────────────────────────
menu_ws = menu_wb['MENU']

thin_border      = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
center_alignment = Alignment(horizontal='center', vertical='center')
header_fill      = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
white_font       = Font(size=18, color="FFFFFF", bold=True)

start_row = 10
start_col = 5

# Build consolidated summary across all model sheets
consolidated = []
required_cols = ['BRAND', 'CPU', 'Form Factor', 'Screen_Size', 'Screen_Type', 'Qty']

for sheet_name in final_wb.sheetnames:
    ws_iter = final_wb[sheet_name]
    data = ws_iter.values
    try:
        columns = next(data)
    except StopIteration:
        continue
    df_sheet = pd.DataFrame(data, columns=columns)
    if all(col in df_sheet.columns for col in required_cols):
        grp = df_sheet.groupby(
            ['BRAND', 'CPU', 'Form Factor', 'Screen_Size', 'Screen_Type'], dropna=False
        ).agg({'Qty': 'sum'}).reset_index()
        grp.insert(0, 'Model Name', sheet_name)
        consolidated.append(grp)

if consolidated:
    final_df = pd.concat(consolidated, ignore_index=True)
else:
    final_df = pd.DataFrame(columns=['Model Name', 'BRAND', 'CPU', 'Form Factor', 'Screen_Size', 'Screen_Type', 'Qty'])

# Table header banner
menu_ws.merge_cells("E8:K8")
menu_ws["E8"].value = "MASTER SUMMARY TABLE"
menu_ws["E8"].alignment = Alignment(horizontal="center", vertical="center")
menu_ws["E8"].fill = PatternFill(fill_type="solid", start_color="0000FF", end_color="0000FF")
menu_ws["E8"].font = Font(size=18, color="FFFFFF", bold=True)

# Write dataframe
from openpyxl.utils import get_column_letter as _gcl
r0 = start_row
c0 = start_col
for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True)):
    for c_idx, val in enumerate(row):
        cell = menu_ws.cell(row=r0 + r_idx, column=c0 + c_idx, value=val)
        if r_idx == 0:
            cell.font = white_font
            cell.alignment = center_alignment
            cell.fill = header_fill
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border

# Freeze header row
menu_ws.freeze_panes = menu_ws.cell(row=r0 + 1, column=c0)

# Add Excel Table
last_row = r0 + len(final_df)
last_col = c0 + len(final_df.columns) - 1
ref = f"{_gcl(c0)}{r0}:{_gcl(last_col)}{last_row}"

base = "MenuSummaryTable"
existing_tables = set(menu_wb['MENU'].tables.keys()) if hasattr(menu_wb['MENU'], 'tables') else set()
name = base
i = 1
while name in existing_tables:
    i += 1
    name = f"{base}{i}"

tbl = Table(displayName=name, ref=ref)
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True,   showColumnStripes=False,
)
menu_ws.add_table(tbl)

# Column widths for summary table
for i, col_name in enumerate(final_df.columns):
    menu_ws.column_dimensions[_gcl(c0 + i)].width = max(14, min(40, len(str(col_name)) + 15))

print("Summary table built OK")


# ─────────────────────────────────────────────
#  STEP 8 — Save final workbook locally
# ─────────────────────────────────────────────
menu_wb.save(LOCAL_OUTPUT)
print(f"Final workbook saved locally: {LOCAL_OUTPUT}")


# ─────────────────────────────────────────────
#  STEP 9 — Upload finished file back to Dropbox
# ─────────────────────────────────────────────
print(f"Uploading result to Dropbox: {OUTPUT_PATH}")
with open(LOCAL_OUTPUT, "rb") as f:
    file_bytes = f.read()

dbx.files_upload(
    file_bytes,
    OUTPUT_PATH,
    mode=dropbox.files.WriteMode.overwrite
)
print(f"Upload complete: {OUTPUT_PATH}")
print(f"Done! Processed {total_lines} rows → {OUTPUT_FILENAME}")
