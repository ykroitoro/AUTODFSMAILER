import pandas as pd
import os
import openpyxl
import xlsxwriter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
#from UliPlot.XLSX import auto_adjust_xlsx_column_width
from typing import NoReturn


#clear shell windows
os.system('cls' if os.name == 'nt' else 'clear')


# Read the data from the DELLOUTLET.XLSX workbook
df = pd.read_excel('DELLOUTLET.XLSX')

# Rename the 'OUTLET Sku' column to 'Outlet SKU'
#df = df.rename(columns={'OUTLET Sku': 'Outlet SKU'})

# Group the data by 'Outlet SKU'
grouped = df.groupby('Outlet SKU')

# Summarize the data by adding the total number of related 'Service Tag' and 
# calculating the average of the 'Reseller Price' column
result = grouped['Service Tag'].count().reset_index()
result['Reseller Price '] = grouped['Reseller Price '].mean().reset_index()['Reseller Price ']

# Add all the columns from the original file to the summarized data
result = pd.merge(result, df, on='Outlet SKU', how='left')

# Drop duplicates from the summarized data
result = result.drop_duplicates(subset='Outlet SKU')

# Rename the 'Service Tag_x' column to 'Qty'
result = result.rename(columns={'Outlet SKU': 'SKU#'})

# Rename the 'Service Tag_x' column to 'Qty'
result = result.rename(columns={'Service Tag_x': 'Qty'})

# Rename the 'Reseller Price_x' column to 'Cost'
result = result.rename(columns={'Reseller Price _x': 'Cost'})

# Rename the 'Product' column to 'MPN'
result = result.rename(columns={'Product': 'MPN'})

# Rename the 'Processor' column to 'CPU'
result = result.rename(columns={'Processor': 'CPU'})

# Rename the 'Graphics' column to 'VIDEO CARD'
result = result.rename(columns={'Graphics': 'VIDEO CARD'})

# Rename the 'Media Bay' column to 'ODD'
result = result.rename(columns={'Media Bay': 'ODD'})

# Rename the 'Mem Total' column to 'MEM'
result = result.rename(columns={'Mem Total': 'MEM'})

# Rename the 'Default Warranty' column to 'WARRANTY'
result = result.rename(columns={'Default Warranty': 'WARRANTY'})

# Rename the 'Display' column to 'DISPLAY'
result = result.rename(columns={'Display': 'DISPLAY'})

# Drop the 'Service Tag_y' and 'Reseller Price_y' columns
result = result.drop(columns=['Service Tag_y', 'Reseller Price _y','Condition','Prc. Qty','Networking','Chassis','Battery','Camera','Keyboard','Mouse / Touchpad','Color','Power','Feature 1','Feature 2','Feature 3','Feature 4','Feature 5','HARDWARE_UPGRADE','FAMILY_NAME','FAMILY_NAME','LOB','PRICE_NEW','FAMILY_SERIES'])

# Calculate the 'Price' column with the formula round(round(cost/.85,0)/5,0)*5
result['Price'] = round(round(result['Cost']/.85,0)/5,0)*5

# Rename the 'Price' column to 'PRICE'
result = result.rename(columns={'Price': 'PRICE'})


# Write the summarized data to a new Excel file
result.to_excel('DELL_LIST_GPT.xlsx', index=False)

#OPEN THE NEW WORKBOOK AND REARRANGE THE COLUMNS

# Read the excel file
df = pd.read_excel("DELL_LIST_GPT.xlsx")

# Drop the 'Cost' column
df = df.drop(columns=["Cost"])

# Rearrange the columns in the specified order
df = df[['Segment', 'SKU#', 'MPN', 'Qty', 'PRICE', 'CPU', 'MEM', 'HDD', 'ODD', 'OS', 'DISPLAY', 'VIDEO CARD', 'WARRANTY']]


# Remove the word "Outlet" from the "MPN" column
df["MPN"] = df["MPN"].str.replace("Outlet", "")

# Write the dataframe to an excel file
df.to_excel("DELL_LIST.xlsx", index=False)




# PART 2 - FORMATING DATA INTO A NEW EXCEL WORKBOOK WITH MULTIPLE TABS, FOR EACH SEGMENT.


import pandas as PD

#CREATING MASTER EXCEL WITH MULTIPLE TAB FOR EACH CATEGORY


# current dateTime
now = datetime.now()
# convert to string
date_time_str = now.strftime("%m-%d-%y %H-%M-%S")
#print('DateTime String:', date_time_str)

df=PD.read_excel('DELL_LIST.xlsx')
#df.head()
# set currency format variable

filename = ('DELLTODAY.xlsx')

# INSERT INFORMATION PAGE



df["Segment"].unique()

for segtype in df['Segment'].unique():

    sheetread=(segtype + ' - READ')
    print (sheetread)

print('')
print('--------------------')
print('')               

writer = PD.ExcelWriter(filename,engine = 'xlsxwriter')

for segtype in df['Segment'].unique():
    newDF = df[df['Segment'] == segtype]
    newDFSort = newDF.sort_values('MPN') #sort by part number
    newDFSort.to_excel(writer,sheet_name = segtype, index = False) # write a sheet with all the segment data CREATING A SHEET WITH A MAX OF 25 CHARACTERS IN THE NAME 
                           
    



writer.close()



# FORMATING THE EXCEL TO CREATE TABLES IN EACH WORKSHEET WITH FORMATED LOOK AND FILTERS

from openpyxl import load_workbook
wb = load_workbook(filename)

sheets = wb.sheetnames
i=1
for s_name in sheets:
    #print(s_name)
    # strip blank spaces from table name
    tablename = s_name.replace(' ','')
    sheet=wb[s_name]
    
    
    # Get the dimensions of the dataframe.
    (max_row, max_col) = df.shape
    #location of last cell in the current sheet
    lastcell = xlsxwriter.utility.xl_col_to_name(max_col)+str(max_row)
    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df.columns]
    # Add the Excel table structure. Pandas will add the data.
    table = Table(displayName = tablename, ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

    #CHANGE PRICE ROW HEADER
    sheet['E1']='PRICE'
    #FORMAT COLUMN E - PRICE COLUMN AS USD
    s_max = sheet.max_row
    for row in range(2,s_max+1):
        s_row = ("E"+ str(row))
        sheet[("E" + str(row))].number_format = '$#,##0_-'

   
    # Make the columns wider for clarity.
    #Autofit column in active sheet
    
    # Auto-adjust columns' width

    def autofit_columns(worksheet):
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length + 2

    #sheet.column_dimensions['A'].width = 25
    #sheet.column_dimensions['B'].width = 25
    #sheet.column_dimensions['C'].width = 35
    #sheet.column_dimensions['D'].width = 10
    #sheet.column_dimensions['E'].width = 10
    #sheet.column_dimensions['F'].width = 75
    #sheet.column_dimensions['G'].width = 20
    #sheet.column_dimensions['H'].width = 20
    #sheet.column_dimensions['I'].width = 30
    #sheet.column_dimensions['J'].width = 20
    #sheet.column_dimensions['K'].width = 75
    #sheet.column_dimensions['L'].width = 100

    donesheet = (s_name+'- Processed')
    print(donesheet)

print('')
print('DONE ....')               

# INSERT INFORMATION IMAGE
# INSERT INFORMATION SHEET IMAGE
sheet2 = wb.create_sheet(title="INFO")

#sheet=wb['INFORMATION']


sheet = wb.copy_worksheet(wb["INFO"])
sheet.title = "INFORMATION"
wb.move_sheet("INFORMATION", -(len(wb.sheetnames)-1))
img = openpyxl.drawing.image.Image('INFORMATION.jpg')
img.anchor = 'A1'
sheet.add_image(img)

del wb['INFO']


wb.save('/Users/yosi/Desktop/DELLPROJECT/LISTS/DELL_FACTORY_CERTIFIED_'+date_time_str+'.xlsx')



#DELETING UNNECESSARY FILES
file_path = "DELL_LIST.XLSX"

if os.path.exists(file_path):
    os.remove(file_path)

file_path = "DELL_LIST_GPT.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)


file_path = "DELLTODAY.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)


