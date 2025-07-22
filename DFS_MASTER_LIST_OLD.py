import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Define file path
file_path = 'DFS_LIST.xlsx'

# Load the Excel workbook
df = pd.read_excel(file_path, sheet_name='Worksheet')

# Create 'Model Name' column
df['Model Name'] = df['Model'].apply(lambda x: x.split()[1] if len(x.split()) > 1 else '')
df.insert(df.columns.get_loc('Model') + 1, 'Model Name', df.pop('Model Name'))

# Create 'Model Number' column
df['Model Number'] = df['Model'].apply(lambda x: x.split()[2] if len(x.split()) > 2 and x.split()[2].isdigit() else '')
df.insert(df.columns.get_loc('Model Name') + 1, 'Model Number', df.pop('Model Number'))

# Create 'Form Factor' column
df['Form Factor'] = df['Model'].apply(lambda x: next((factor for factor in ['SFF', 'MFF', 'MT', 'Tower', 'AIO', 'NB'] if factor in x), ''))
df.insert(df.columns.get_loc('Model Number') + 1, 'Form Factor', df.pop('Form Factor'))

# Create 'CPU' column
df['CPU'] = df['Processor'].apply(lambda x: x[14:22] if isinstance(x, str) else '')
df['RAM'] = df['RAM'].apply(lambda x: x[:2] if isinstance(x, str) else x)

# Create 'ODD' column
df['ODD'] = df['Description'].apply(lambda x: 'DVDRW' if isinstance(x, str) and 'DVDRW' in x else ('DVD' if isinstance(x, str) and 'DVD' in x else ''))
df.insert(df.columns.get_loc('Form Factor') + 1, 'ODD', df.pop('ODD'))

# Create 'HDD' column
df['HDD'] = df['Description'].apply(lambda x: next((size for size in ['128GB', '256GB', '500GB', '512GB', '250GB', '320GB', '1000GB', '1024GB', '768GB'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('ODD') + 1, 'HDD', df.pop('HDD'))


# Create 'Screen' column
df['Screen'] = df['Description'].apply(lambda x: next((size for size in ['14-in', '13.3-in', '15.6-in','15,2','13,3','10-in','12.5-in','16-in','11.6-in','2N1','17.3'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('HDD') + 1, 'Screen', df.pop('Screen'))



# Move 'CPU' and 'RAM' right after 'Model Number'
df.insert(df.columns.get_loc('Model Number') + 1, 'CPU', df.pop('CPU'))
df.insert(df.columns.get_loc('Model Number') + 2, 'RAM', df.pop('RAM'))

# Drop 'Processor' column
df.drop(columns=['Processor'], inplace=True)

# Sort the dataframe
df.sort_values(by=['Model Name', 'Model Number', 'Form Factor', 'CPU', 'RAM', 'HDD'], inplace=True)

# Create 'LISTS' directory if it doesn't exist
if not os.path.exists('LISTS'):
    os.makedirs('LISTS')

# Save the dataframe to Excel
current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
new_file_path = os.path.join('LISTS', f'DELL_LIST_GRADE_A_{current_time}.xlsx')
df.to_excel(new_file_path, index=False)

# Load the workbook and adjust column widths
wb = load_workbook(new_file_path)
ws = wb.active

# Adjust column widths
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    column_letter = col[0].column_letter
    ws.column_dimensions[column_letter].width = max_length + 2

# Rename columns
ws.cell(row=1, column=4, value='Model Number')  # Column D
ws.cell(row=1, column=12, value='Cost')  # Column K

# Add 'Price' column and calculate (column K / 0.75) + 10, formatted as currency
col_idx_k = 12
ws.insert_cols(col_idx_k + 1)
ws.cell(row=1, column=col_idx_k + 1, value='Price')

for row in range(2, ws.max_row + 1):
    cost = ws.cell(row=row, column=col_idx_k).value
    if isinstance(cost, (int, float)):
        price = round(((cost / 0.75) + 10)/5,0)*5
        price_cell = ws.cell(row=row, column=col_idx_k + 1, value=price)
        price_cell.number_format = '$#,##0.00'

# Delete 'Cost' and 'Product ID' 
cost_col_idx = 12
product_id_col_idx = 1  # Assuming 'Product ID' is the first column
ws.delete_cols(cost_col_idx)
ws.delete_cols(product_id_col_idx)

# Make J1 bold and center
cell_j1 = ws['J1']
cell_j1.font = Font(bold=True)
cell_j1.alignment = Alignment(horizontal='center')

# Save the intermediate workbook
wb.save(new_file_path)

# Re-load workbook to create tabs for each model name

# Load the Excel workbook
df = pd.read_excel(new_file_path, sheet_name='Sheet1')



# wb = load_workbook(new_file_path)
header = list(df.columns)



for model in df['Model Name'].unique():
    model_df = df[df['Model Name'] == model]
    model_ws = wb.create_sheet(title=model)
    # Write the header
    for col_idx, col_name in enumerate(header, 1):
        cell = model_ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    # Write the data
    for r_idx, row in enumerate(model_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            model_ws.cell(row=r_idx, column=c_idx, value=value)
    # Adjust column widths for the new sheet
    for col in model_ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        if max_length > 0:
            column_letter = col[0].column_letter
            model_ws.column_dimensions[column_letter].width = max_length + 2
    # Format column K as currency
    for row in range(2, model_ws.max_row + 1):
        model_ws.cell(row=row, column=11).number_format = '$#,##0.00'
    # Delete column A
    model_ws.delete_cols(1)
   
    
    # Autofit all columns
    for col in model_ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        column_letter = col[0].column_letter
        if max_length > 0:
            model_ws.column_dimensions[column_letter].width = max_length + 2

    # Set column J-I-K width
    model_ws.column_dimensions['J'].width = 10
    model_ws.column_dimensions['L'].width = 120
    model_ws.column_dimensions['k'].width = 20



# Delete the original worksheet
wb.remove(wb['Sheet1'])

# Delete any other tabs that are not "Latitude", "OptiPlex", or "Precision"
for sheet_name in wb.sheetnames:
    if sheet_name not in ["Latitude", "OptiPlex", "Precision"]:
        wb.remove(wb[sheet_name])

# Save the final workbook
wb.save(new_file_path)

# Output the message with the total number of lines
total_lines = len(df)  # Total number of lines in the original dataframe
print(f"Done Processing... Total number of lines: {total_lines}")
