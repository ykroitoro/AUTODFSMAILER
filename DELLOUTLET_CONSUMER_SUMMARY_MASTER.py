import pandas as pd
import numpy as np
import os
import openpyxl
import xlsxwriter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime

#current file path
path = "/Users/yosi/Desktop/DELLPROJECT/LISTS"

# Load the Excel file into a pandas dataframe
df = pd.read_excel("Consumer List.xlsx")

# Delete the first 4 columns
df = df.drop(df.columns[:4], axis=1)

# Change the column name "Dell Outlet Product" to "Dell Product"
df = df.rename(columns={"Dell Outlet Product": "Dell Product"})

# Insert a column "Family" before the column "Dell Product"
df.insert(df.columns.get_loc("Dell Product"), "Family", value=None)

# Set the value of the "Family" column to the first word found in the "Dell Product" column
df["Family"] = df["Dell Product"].apply(lambda x: x.split()[0])

# Change the column name "Outlet SKU" to "Dell SKU"
#df = df.rename(columns={"Outlet SKU": "Dell SKU"})

# Change the column name "Outlet List Price" to "Price"
df = df.rename(columns={"Outlet List Price": "Price"})

# Change the column name "Reseller Price" to "Cost"
df = df.rename(columns={"Reseller Price": "Cost"})

# Apply the formula to the "Price" column
df["Price"] = np.round(np.round(df["Cost"] / 0.85, 0) / 5, 0) * 5

# Remove the "Cost" column
df = df.drop(columns=["Cost"])

# Move the "Units" column 3 places to the right
column_list = list(df.columns)
column_list.remove("Units")
column_list = column_list[:column_list.index("Family")+2] + ["Units"] + column_list[column_list.index("Family")+1:]

# Move the "Price" column one place after the "Units" column
column_list.remove("Price")
column_list = column_list[:column_list.index("Units")+1] + ["Price"] + column_list[column_list.index("Units")+1:]

# Reorder the columns
df = df[column_list]

# Save the modified dataframe to a new Excel file with the desired name
df.to_excel("Dell Consumer List.xlsx", index=False)




#CREATING MASTER EXCEL WITH MULTIPLE TAB FOR EACH CATEGORY

import pandas as PD
# current dateTime
now = datetime.now()
# convert to string
date_time_str = now.strftime("%m-%d-%y %H-%M-%S")
#print('DateTime String:', date_time_str)

df=PD.read_excel('Dell Consumer List.xlsx')

filename = ('DELLCONSUMER.xlsx')


df["Family"].unique()

for segtype in df['Family'].unique():

    sheetread=(segtype + ' - READ')
    print (sheetread)

print('')
print('--------------------')
print('')               

writer = PD.ExcelWriter(filename,engine = 'xlsxwriter')

for segtype in df['Family'].unique():
    newDF = df[df['Family'] == segtype]
    newDFSort = newDF.sort_values('Dell Product') #sort by part number
    newDFSort.to_excel(writer,sheet_name = segtype, index = False) # write a sheet with all the segment data CREATING A SHEET WITH A MAX OF 25 CHARACTERS IN THE NAME 
                           
    



writer.close()



# FORMATING THE EXCEL TO CREATE TABLES IN EACH WORKSHEET WITH FORMATED LOOK AND FILTERS

from openpyxl import load_workbook
wb = load_workbook(filename)

sheets = wb.sheetnames
i=1
for s_name in sheets:
    #print(s_name)
    # strip blank spaces from table name
    tablename = s_name.replace(' ','')
    sheet=wb[s_name]
    
    
    # Get the dimensions of the dataframe.
    (max_row, max_col) = df.shape
    #location of last cell in the current sheet
    lastcell = xlsxwriter.utility.xl_col_to_name(max_col)+str(max_row)
    # Create a list of column headers, to use in add_table().
    column_settings = [{'header': column} for column in df.columns]
    # Add the Excel table structure. Pandas will add the data.
    table = Table(displayName = tablename, ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    sheet.add_table(table)

    #CHANGE PRICE ROW HEADER
    sheet['E1']='PRICE'
    #FORMAT COLUMN E - PRICE COLUMN AS USD
    s_max = sheet.max_row
    for row in range(2,s_max+1):
        s_row = ("E"+ str(row))
        sheet[("E" + str(row))].number_format = '$#,##0_-'

   
    # Make the columns wider for clarity.
    #Autofit column in active sheet
    
    # Auto-adjust columns' width
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 35
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 75
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 75
    sheet.column_dimensions['L'].width = 100

    donesheet = (s_name+'- Processed')
    print(donesheet)

print('')
print('DONE ....')               

# INSERT INFORMATION IMAGE
# INSERT INFORMATION SHEET IMAGE
sheet2 = wb.create_sheet(title="INFO")

#sheet=wb['INFORMATION']


sheet = wb.copy_worksheet(wb["INFO"])
sheet.title = "INFORMATION"
wb.move_sheet("INFORMATION", -(len(wb.sheetnames)-1))
img = openpyxl.drawing.image.Image('INFORMATION.jpg')
img.anchor = 'A1'
sheet.add_image(img)

del wb['INFO']

#SAVING FINAL FILE
wb.save('/Users/yosi/Desktop/DELLPROJECT/LISTS/DELL_CONSUMER_FACTORY_CERTIFIED_'+date_time_str+'.xlsx')


#DELETING UNNECESSARY FILES
file_path = "Dell Consumer List.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)

file_path = "DELLCONSUMER.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)

