import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import copy
from colorama import Fore, Style, init
import tkinter as tk
from tkinter import filedialog, messagebox
import subprocess

# Initialize colorama
init(autoreset=True)

# Function to process the script
def process_script(user_choice, delete_columns):
    
    try:
        # Select the file
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            messagebox.showwarning("Warning", "No file selected. Process canceled.")
            return
        
        # Load the Excel workbook
        df = pd.read_excel(file_path, sheet_name='Worksheet')
        
        # Create 'Model Name' column
        df['Model Name'] = df['Model'].apply(lambda x: x.split()[1] if len(x.split()) > 1 else '')
        df.insert(df.columns.get_loc('Model') + 1, 'Model Name', df.pop('Model Name'))

        # Create 'Model Number' column
        df['Model Number'] = df['Model'].apply(lambda x: x.split()[2] if len(x.split()) > 2 and x.split()[2].isdigit() else (x.split()[2] if len(x.split()) > 2 else ''))
        df.insert(df.columns.get_loc('Model Name') + 1, 'Model Number', df.pop('Model Number'))

        # Create 'Form Factor' column
        df['Form Factor'] = df['Description'].apply(lambda x: next((factor for factor in ['SFF', 'MFF', 'MT', 'Tower', 'AIO', 'NB'] if isinstance(x, str) and factor in x), 'NB'))
        df.insert(df.columns.get_loc('Model Number') + 1, 'Form Factor', df.pop('Form Factor'))

        # Create 'CPU' column
        df['CPU'] = df['Processor'].apply(lambda x: x[14:22] if isinstance(x, str) else '')
        df['RAM'] = df['RAM'].apply(lambda x: x[:2] if isinstance(x, str) else x)

        # Create 'ODD' column
        df['ODD'] = df['Description'].apply(lambda x: 'DVDRW' if isinstance(x, str) and 'DVDRW' in x else ('DVD' if isinstance(x, str) and 'DVD' in x else ''))
        df.insert(df.columns.get_loc('Form Factor') + 1, 'ODD', df.pop('ODD'))

        # Create 'HDD' column
        df['HDD'] = df['Description'].apply(lambda x: next((size for size in ['128GB', '256GB', '500GB', '512GB', '250GB', '320GB', '1000GB', '1024GB', '768GB'] if isinstance(x, str) and size in x), ''))
        df.insert(df.columns.get_loc('ODD') + 1, 'HDD', df.pop('HDD'))

        # Create 'Screen Size' column
        df['Screen_Size'] = df['Description'].apply(lambda x: next((size for size in ['14-in', '13.3-in', '15.6-in','15,2','13,3','10-in','12.5-in','16-in','11.6-in','2N1','17.3','19.5-in', 'TAB', '24-in'] if isinstance(x, str) and size in x), ''))
        df.insert(df.columns.get_loc('HDD') + 1, 'Screen_Size', df.pop('Screen_Size'))

        # Create 'Screen Type' column
        df['Screen_Type'] = df['Description'].apply(lambda x: next((size for size in ['NoTCH', 'TCH'] if isinstance(x, str) and size in x), ''))
        df.insert(df.columns.get_loc('Screen_Size') + 1, 'Screen_Type', df.pop('Screen_Type'))

        # Create 'CAMERA' column
        df['Camera'] = df['Description'].apply(lambda x: next((size for size in ['NoCAM','CAM'] if isinstance(x, str) and size in x), ''))
        df.insert(df.columns.get_loc('Screen_Type') + 1, 'Camera', df.pop('Camera'))

        # Move 'CPU' and 'RAM' right after 'Model Number'
        df.insert(df.columns.get_loc('Model Number') + 1, 'CPU', df.pop('CPU'))
        df.insert(df.columns.get_loc('Model Number') + 2, 'RAM', df.pop('RAM'))

        # Drop 'Processor' column
        df.drop(columns=['Processor'], inplace=True)

        # Sort the dataframe
        df.sort_values(by=['Model Name', 'Model Number', 'Form Factor', 'CPU', 'RAM', 'HDD'], inplace=True)

        # Create 'LISTS' directory if it doesn't exist
        if not os.path.exists('LISTS'):
            os.makedirs('LISTS')

        # Save the dataframe to Excel
        current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        new_file_path = os.path.join('LISTS', f'DELL_LIST_GRADE_A_{current_time}.xlsx')
        df.to_excel(new_file_path, index=False)
        


    # Load the workbook and adjust column widths
        wb = load_workbook(new_file_path)
        ws = wb.active

        GREEN = '\033[92m'
        RESET = '\033[0m'

        # Add Prompt to calculate the 'Price' column
        # user_choice = input(GREEN + "Do you want to calculate the Price column? (yes/no): " + RESET).strip().lower()  # Add Prompt to calculate the 'Price' column
        # delete_columns = input(GREEN + "Do you want to delete both 'Cost' and 'Product ID' columns? (yes/no): " + RESET).strip().lower()

        # Adjust column widths
        for col in ws.columns:
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            column_letter = col[0].column_letter
            ws.column_dimensions[column_letter].width = max_length + 2

        # Rename columns
        ws.cell(row=1, column=4, value='Model Number')  # Column D
        ws.cell(row=1, column=14, value='Cost')  # Column M

        # Add 'Price' column and calculate (column K / 0.75) + 10, formatted as currency

        col_idx_k = 14
        ws.insert_cols(col_idx_k + 1)


        # assigning correct header to price or bid column
        if user_choice == 'yes':
            ws.cell(row=1, column=col_idx_k + 1, value='Price')
        else:
                ws.cell(row=1, column=col_idx_k + 1, value='Bid Price')


        if user_choice == 'yes':
            # Placeholder logic for calculating the Price (you can replace this with actual logic)

    
            for row in range(2, ws.max_row + 1):
                cost = ws.cell(row=row, column=col_idx_k).value
                if isinstance(cost, (int, float)):
                    price = round(((cost / 0.75) + 10)/5,0)*5
                    price_cell = ws.cell(row=row, column=col_idx_k + 1, value=price)
                    price_cell.number_format = '$#,##0.00'

        else:
            # Leave the 'Price' column blank
            for row in range(2, ws.max_row + 1):
                price = ''
            
  

        # MAKE EVERY LINE IN OS COLUMN = TO WINDOWS 10 PROFESSIONAL
        col_idx_k = 16
        # ws.insert_cols(col_idx_k + 1)
        # ws.cell(row=1, column=col_idx_k + 1, value='Price')

        for row in range(2, ws.max_row + 1):
            OS = ws.cell(row=row, column=col_idx_k).value
            if OS == 'No Operating System':
                OSN = 'Windows 10 Professional'
                ws.cell(row=row, column=col_idx_k).value=OSN
       

        # MAKE EVERY LINE IN OPTIPLEX SHEET SCREEN TYPE = 'NO MONITOR'
        col_idx_k = 10
        # ws.insert_cols(col_idx_k + 1)
        # ws.cell(row=1, column=col_idx_k + 1, value='Price')

        for row in range(2, ws.max_row + 1):
            SCRN_S = ws.cell(row=row, column=col_idx_k).value
            if SCRN_S is None or SCRN_S == "":
                SCRN = 'No Monitor'
                ws.cell(row=row, column=col_idx_k).value=SCRN



        # Delete 'Cost' and 'Product ID' 
        cost_col_idx = 14
        product_id_col_idx = 1  # Assuming 'Product ID' is the first column
        # ws.delete_cols(cost_col_idx)
        # ws.delete_cols(product_id_col_idx)


        if delete_columns == 'yes':
            # Assuming 'df' is the DataFrame being manipulated
            ws.delete_cols(cost_col_idx)
            ws.delete_cols(product_id_col_idx)
            print("'Cost' and 'Product ID' columns have been deleted.")
        else:
            print("Columns 'Cost' and 'Product ID' will be retained.")
    


        # Make J1 bold and center
        cell_j1 = ws['J1']
        cell_j1.font = Font(bold=True)
        cell_j1.alignment = Alignment(horizontal='center')

        # Save the intermediate workbook
        wb.save(new_file_path)

        # Re-load workbook to create tabs for each model name

        # Load the Excel workbook
        df = pd.read_excel(new_file_path, sheet_name='Sheet1')



        # wb = load_workbook(new_file_path)
        header = list(df.columns)



        for model in df['Model Name'].unique():
            model_df = df[df['Model Name'] == model]
            model_ws = wb.create_sheet(title=model)
            # Write the header
            for col_idx, col_name in enumerate(header, 1):
                cell = model_ws.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            # Write the data
            for r_idx, row in enumerate(model_df.values, 2):
                for c_idx, value in enumerate(row, 1):
                    model_ws.cell(row=r_idx, column=c_idx, value=value)
            # Adjust column widths for the new sheet
            for col in model_ws.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                if max_length > 0:
                    column_letter = col[0].column_letter
                    model_ws.column_dimensions[column_letter].width = max_length + 2
            # Format PRICE COLUMN as currency

            if delete_columns == 'yes':
                for row in range(2, model_ws.max_row + 1):
                    model_ws.cell(row=row, column=13).number_format = '$#,##0.00'
            
            else:
                for row in range(2, model_ws.max_row + 1):
                    model_ws.cell(row=row, column=14).number_format = '$#,##0.00'
            
                for row in range(2, model_ws.max_row + 1):    
                    model_ws.cell(row=row, column=15).number_format = '$#,##0.00'
            
            # Delete column A
            # model_ws.delete_cols(1)
   
    
            # Autofit all columns
            for col in model_ws.columns:
                max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                column_letter = col[0].column_letter
                if max_length > 0:
                    model_ws.column_dimensions[column_letter].width = max_length + 2

        if delete_columns == 'yes':
            # Set column J-I-K width
            model_ws.column_dimensions['L'].width = 15
            model_ws.column_dimensions['N'].width = 120
            model_ws.column_dimensions['M'].width = 30
    
        else:
            # Set column J-I-K width
            model_ws.column_dimensions['L'].width = 15
            model_ws.column_dimensions['N'].width = 30
            model_ws.column_dimensions['M'].width = 15
            model_ws.column_dimensions['O'].width = 120

            # APPLY FILTERS
            model_ws.auto_filter.ref = ws.dimensions  # Applies filter to the range of dataws.auto_filter.ref = ws.dimensions  # Applies filter to the range of data


        # Delete the original worksheet
        wb.remove(wb['Sheet1'])

        # Delete any other tabs that are not "Latitude", "OptiPlex", or "Precision"
        for sheet_name in wb.sheetnames:
            if sheet_name not in ["Latitude", "OptiPlex", "Precision"]:
                wb.remove(wb[sheet_name])




    
        # Save the final workbook
        wb.save(new_file_path)

        #********** merging all worksheets into the menu workbook ***********

        # Define file paths
        final_file_path = new_file_path  # Recently created workbook
        menu_file_path = 'DELL_LIST_MENU.xlsx'  # Workbook with the "Menu" sheet

        # --- FUNCTIONALITY TO MERGE BOTH WORKBOOKS ---

        # Load the final workbook 'DFS_LIST_PROCESSED.xlsx' (the one with new sheets)
        final_wb = load_workbook(final_file_path)

        # Load the 'DELL_LIST_MENU.xlsx' workbook (which contains the Menu sheet)
        menu_wb = load_workbook(menu_file_path)

        # Copy each sheet from 'DFS_LIST_PROCESSED.xlsx' into 'DELL_LIST_MENU.xlsx'
        for sheet_name in final_wb.sheetnames:
            # Get the current sheet from the final workbook
            source_sheet = final_wb[sheet_name]

            # If the sheet already exists in the menu workbook, remove it to avoid duplication
            if sheet_name in menu_wb.sheetnames:
                del menu_wb[sheet_name]

            # Create a new sheet in 'DELL_LIST_MENU.xlsx' with the same name as in 'DFS_LIST_PROCESSED.xlsx'
            target_sheet = menu_wb.create_sheet(sheet_name)

            # Copy each cell from the source sheet to the new target sheet, preserving values and styles
            for row in source_sheet.iter_rows():
                for cell in row:
                    new_cell = target_sheet[cell.coordinate]
                    new_cell.value = cell.value
            
                     # Copy cell font, fill, alignment, number_format, protection, and border
                    if cell.has_style:
                        new_cell.font = copy.copy(cell.font)
                        new_cell.fill = copy.copy(cell.fill)
                        new_cell.alignment = copy.copy(cell.alignment)
                        new_cell.number_format = cell.number_format  # This is a simple string, no need to copy
                        new_cell.protection = copy.copy(cell.protection)
                        new_cell.border = copy.copy(cell.border)

            # Copy dimensions (column width, row height)
            for col in source_sheet.column_dimensions:
                target_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]

            for row in source_sheet.row_dimensions:
                target_sheet.row_dimensions[row] = source_sheet.row_dimensions[row]

            # Copy merged cells
            for merged_range in source_sheet.merged_cells.ranges:
                target_sheet.merge_cells(str(merged_range))

         # --- Add Autofilter to the new sheet ---
            # Assuming the first row contains headers and we want to apply the autofilter over all data
            max_column = target_sheet.max_column  # Get the last column number with data
            max_row = target_sheet.max_row  # Get the last row number with data
            target_sheet.auto_filter.ref = f"A1:{target_sheet.cell(row=1, column=max_column).coordinate}"
            # The range is from A1 to the last column in the first row (header row)

        # ADDING HYPER LINKS TO EACH TAB IN THE MENU

        # Select the 'Menu' worksheet
        menu_ws = menu_wb['MENU']

        # --- Add hyperlinks under each image ---

        # Add hyperlink for the Latitude tab under the first image
        menu_ws['N8'] = 'Go to Latitude'
        menu_ws['N8'].hyperlink = '#Latitude!A1'  # Hyperlink to the "Latitude" tab
        menu_ws['N8'].font = Font(color="0000FF", underline="single", size=16)  # Style it like a hyperlink

        # Add hyperlink for the OptiPlex tab under the second image
        menu_ws['N10'] = 'Go to OptiPlex'
        menu_ws['N10'].hyperlink = '#OptiPlex!A1'  # Hyperlink to the "OptiPlex" tab
        menu_ws['N10'].font = Font(color="0000FF", underline="single", size=16)

        # Add hyperlink for the Precision tab under the third image
        menu_ws['N12'] = 'Go to Precision'
        menu_ws['N12'].hyperlink = '#Precision!A1'  # Hyperlink to the "Precision" tab
        menu_ws['N12'].font = Font(color="0000FF", underline="single", size=16)

        # CREATING SUMMARY TABLES FOR EACH TAB IN MENU SHEET
        # Set the starting row for summary tables


        # CREATING SUMMARY TABLES
        # Set the starting row and column for the summary tables on the MENU tab
        start_row = 17
        start_col = 5  # Start from column A

        # Define a border style for the tables
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Define center alignment
        center_alignment = Alignment(horizontal='center', vertical='center')

        # Define background fills for headers
        blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
        orange_fill = PatternFill(start_color="CC5500", end_color="CC5500", fill_type="solid")
        green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
        black_fill = PatternFill(start_color="000000", end_color="008000", fill_type="solid")

        # Define a white font for the headers
        white_font = Font(color="FFFFFF", bold=True)

        # Function to write a summary table to the MENU tab and adjust column width
        def write_summary_to_menu(ws, df, sheet_name, start_row, start_col):
            print(f"Writing summary for {sheet_name} starting at row {start_row}, column {start_col}")



            # ###Add a total row to the DataFrame, summing the "Qty" column (assuming it exists)
            if "Qty" in df.columns:
                total_units = df["Qty"].sum()
                total_row = pd.DataFrame([["Total", "", total_units, ""]], columns=df.columns)
                df = pd.concat([df, total_row], ignore_index=True)
    
            # Write the sheet name as a header, centered above the table
            header_cell = ws.cell(row=start_row, column=start_col, value=sheet_name)
            header_cell.font = Font(bold=True, size=14, color="FFFFFF")  # Set font size to 14 and bold
            header_cell.alignment = center_alignment
            header_cell.fill = black_fill
            #header_cell.font = white_font

            # Merge cells to center the header over summary columns
            ws.merge_cells(start_row=15, start_column=5, end_row=15, end_column=20)
            ws.cell(15,5, value ="SUMMARY TABLES BY CATEGORY, PROCESSOR AND FORM FACTOR")
            ws.cell(15,5).font = Font(bold=True, size=20, color="FFFFFF")  # Set font size to 14 and bold
            ws.cell(15,5).fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            ws.cell(15,5).alignment = center_alignment


            # Merge cells to center the header over the table columns
            ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(df.columns) - 1)
    
            current_row = start_row + 1

            # Write the dataframe to the MENU tab, starting at the given column
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
                for c_idx, value in enumerate(row, start_col):
                    cell = ws.cell(row=current_row + r_idx, column=c_idx, value=value)
                    # Apply bold font and custom background fills for header row
                    if r_idx == 0:
                        cell.font = white_font  # Set white font
                        # Apply background colors based on the column
                        if c_idx == start_col:
                            cell.fill = blue_fill  # First header
                        elif c_idx == start_col + 1:
                            cell.fill = orange_fill  # Second header
                        elif c_idx == start_col + 2:
                            cell.fill = green_fill  # Third header
                        elif c_idx == start_col + 3:
                            cell.fill = blue_fill  # Fourth header
                    
                        # Apply border and center alignment to all cells
                    cell.border = thin_border
                    cell.alignment = center_alignment

            # Adjust column widths for the table (setting width to 12)
            for col_idx in range(start_col, start_col + len(df.columns)):
                ws.column_dimensions[get_column_letter(col_idx)].width = 12

            # Return the next available column after writing the table
            return start_col + len(df.columns) + 2  # Adding 2 columns of space between tables

 
    

        # Process each worksheet except the 'MENU' worksheet
        for sheet_name in wb.sheetnames:
            if sheet_name != 'MENU':
                ws = wb[sheet_name]
                print(f"Processing sheet: {sheet_name}")
        
                # Load the data into a pandas DataFrame
                data = ws.values
                columns = next(data)[0:]  # First row is the header
                df = pd.DataFrame(data, columns=columns)
        
                # Check if required columns exist for summarization
                if 'CPU' in df.columns and 'Form Factor' in df.columns and 'Qty' in df.columns and 'Screen_Size' in df.columns:
                    print(f"Creating summary for {sheet_name}")

                    # Create summary table grouped by CPU and Form Factor, with total qty
                    summary_df = df.groupby(['CPU', 'Form Factor', 'Screen_Size']).agg({'Qty': 'sum'}).reset_index()

                    # Write the summary table to the MENU tab
                    start_col = write_summary_to_menu(menu_ws, summary_df, sheet_name, start_row, start_col)
                else:
                    print(f"Required columns not found in {sheet_name}")


        # Save the updated 'DELL_LIST_MENU.xlsx' file, which now contains all sheets
        menu_wb.save(new_file_path)


        # Output the message with the total number of lines
        total_lines = len(df)  # Total number of lines in the original dataframe
        print(f"Done Processing... Total number of lines: {total_lines}")







        messagebox.showinfo("Success", f"Process complete. File saved as: {new_file_path}")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")

   # Open the final Excel file
    try:
        subprocess.run(["open", new_file_path])  # Use 'open' for macOS
        print(f"Opened {new_file_path} successfully.")
    except Exception as e:
        print(f"Could not open the file. Error: {e}")

# Create the GUI
def main():
    root = tk.Tk()

    # Add a window icon (for .ico files)
    try:
        root.iconbitmap('MYTECH_LOGO.png')  # Replace with your actual .ico file path
    except Exception as e:
        print(f"Icon not found: {e}")

    # Display the logo image (for .png files)
    from PIL import Image, ImageTk
    logo_image = Image.open('MYTECH_LOGO.ico')  # Replace with your actual image file path
    logo = ImageTk.PhotoImage(logo_image)
    tk.Label(root, image=logo).pack(pady=10)

    root.title("Excel Processor")
    root.geometry("600x400")

    tk.Label(root, text="DELL DFS LIST", font=("Helvetica", 16, "bold")).pack(pady=10)

    tk.Label(root, text="Do you want to calculate the Price column? (yes/no)").pack()
    price_choice_var = tk.StringVar(value="yes")
    tk.Entry(root, textvariable=price_choice_var, width=20).pack(pady=5)

    tk.Label(root, text="Do you want to delete both 'Cost' and 'Product ID' columns? (yes/no)").pack()
    delete_choice_var = tk.StringVar(value="yes")
    tk.Entry(root, textvariable=delete_choice_var, width=20).pack(pady=5)

    

    def run_script():
        user_choice = price_choice_var.get().strip().lower()
        delete_columns = delete_choice_var.get().strip().lower()
        process_script(user_choice, delete_columns)
        root.destroy()

    tk.Button(root, text="Run Script", command=run_script, width=15).pack(pady=10)       
    tk.Button(root, text="Exit", command=root.destroy, width=15).pack(pady=5)

    root.mainloop()

if __name__ == "__main__":
    main()


