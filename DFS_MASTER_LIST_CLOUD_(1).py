# import tkinter as tk
import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
import copy
from dotenv import load_dotenv
from colorama import Fore, Style, init
# from tkinter import filedialog, messagebox
import subprocess
from PIL import Image
import dropbox
from dropbox.files import WriteMode
from dropbox.oauth import DropboxOAuth2FlowNoRedirect
from dropbox.exceptions import AuthError
import io
from io import BytesIO
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.auth.transport.requests import Request
import json




# Initialize colorama
init(autoreset=True)



# Load .env values
load_dotenv()
# print("TENANT_ID:", os.getenv("TENANT_ID"))

SAVE_FOLDER = os.getenv("SAVE_FOLDER", "/tmp")
SAVE_FILENAME = os.getenv("SAVE_PATH", "DFS_LIST.XLSX")
SAVE_PATH = os.path.join(os.getenv("SAVE_FOLDER"), os.getenv("SAVE_FILENAME"))
DROPBOX_ACCESS_TOKEN = os.getenv("DROPBOX_ACCESS_TOKEN")
app_key = os.getenv("DROPBOX_APP_KEY")
app_secret = os.getenv("DROPBOX_APP_SECRET")
refresh_token = os.getenv("DROPBOX_REFRESH_TOKEN")
INPUT_PATH = "/AUTODFSMAILER/DFS_LIST.XLSX"
MENU_TEMPLATE_PATH = "/AUTODFSMAILER/DELL_LIST_MENU.xlsx"
OUTPUT_FILENAME = f"DELL_LIST_GRADE_A_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
# OUTPUT_FILENAME = "DELL_LIST_GRADE_A_LATEST.xlsx"
OUTPUT_PATH = f"/AUTODFSMAILER/LISTS/{OUTPUT_FILENAME}"
# OUTPUT_PATH = os.path.expanduser(os.path.join("~/Dropbox/AUTODFSMAILER/LISTS", OUTPUT_FILENAME))
INPUT_PATH = "/AUTODFSMAILER/DFS_LIST.XLSX"
MENU_TEMPLATE_PATH = "/AUTODFSMAILER/DELL_LIST_MENU.xlsx"
USER_EMAIL = os.getenv("USER_EMAIL")
SUBJECT_KEYWORD = os.getenv("SUBJECT_KEYWORD")
#SAVE_PATH = os.path.join(os.getenv("SAVE_FOLDER", "/tmp"), os.getenv("SAVE_PATH", "DFS_LIST.XLSX"))
SENDER_EMAIL = os.getenv("SENDER_EMAIL")
RECIPIENT_EMAIL = os.getenv("RECIPIENT_EMAIL")
CREDENTIALS_JSON = os.getenv("CREDENTIALS_JSON")
TOKEN_JSON = os.getenv("TOKEN_JSON")
                            

 # Initialize Dropbox client using refresh token (no more token expiration!)
dbx = dropbox.Dropbox(
    app_key=app_key,
    app_secret=app_secret,
    oauth2_refresh_token=refresh_token
)

# dbx = dropbox.Dropbox(DROPBOX_ACCESS_TOKEN)

def sanitize_sheet_title(title):
    if not isinstance(title, str):
        title = str(title) if title is not None else "Unnamed"
    title = title.strip()
    title = title[:31]
    for c in ['\\', '/', '*', '[', ']', ':', '?']:
        title = title.replace(c, '')
    return title or "Unnamed"

### Dropbox input/output paths
##input_path = "/AUTODFSMAILER/DFS_LIST.XLSX"
##menu_template_path = "/AUTODFSMAILER/DELL_LIST_MENU.xlsx"
##timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
##output_filename = f"DELL_LIST_GRADE_A_{timestamp}.xlsx"
##output_path = f"/AUTODFSMAILER/{output_filename}"

# --- Download input file from Dropbox ---
_, input_res = dbx.files_download(INPUT_PATH)
df = pd.read_excel(io.BytesIO(input_res.content), sheet_name="Worksheet")


## upload function for final file version into google drive
print("uploading file to google drive")
SCOPES = ['https://www.googleapis.com/auth/drive.file']

def upload_to_drive_oauth(file_path, file_name):
    creds = None

    # Load token.json content from environment variable
    token_json_content = os.getenv("TOKEN_JSON")
    if token_json_content is None:
        raise ValueError("❌ Environment variable TOKEN_JSON is not set.")

    # Load credentials from the environment variable
    credentials_json = os.getenv("CREDENTIALS_JSON")
    if credentials_json is None:
        raise ValueError("❌ Environment variable CREDENTIALS_JSON is not set.")
    credentials_dict = json.loads(credentials_json)


    # Parse the JSON string into a dictionary
    token_info = json.loads(token_json_content)

    # Create credentials from the token dictionary
    creds = Credentials.from_authorized_user_info(token_info, SCOPES)

    # If no token or expired, login manually
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_config(credentials_dict, SCOPES)
            creds = flow.run_local_server(port=0)

            token_path = os.getenv("TOKEN_JSON")
        if not token_path:
            raise ValueError("Missing TOKEN_JSON path in environment variables")

        with open(token_path, 'w') as token:
            token.write(creds.to_json())

    service = build('drive', 'v3', credentials=creds)

    file_metadata = {'name': file_name}
    media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    file = service.files().create(body=file_metadata, media_body=media, fields='id, webViewLink').execute()
    print(f"File uploaded: {file.get('webViewLink')}")
    return file.get('id')


def main(ans1, ans2):
    
    # Initialize colorama
    init(autoreset=True)
ans1 = "yes"
ans2 = "yes"
    
    # Function to process the script
    #def process_script(ans1, ans2):

print(ans1)
print(ans2)

# Initialize colorama
init(autoreset=True)

# Define file path
file_path = 'DFS_LIST.xlsx'


# Load the Excel workbook
df = pd.read_excel(file_path, sheet_name='Worksheet')
df.fillna('', inplace=True)

# Define values to remove
remove_values = ['TEST30', 'chargeTestFinale']

# Filter out the unwanted rows (only if any of them are found)
found_to_remove = df['Model'].isin(remove_values).any()

if found_to_remove:
    df = df[~df['Model'].isin(remove_values)]
    print("Removed rows with Model values: TEST30 and/or chargeTestFinale")
else:
    print("No rows found with Model values: TEST30 or chargeTestFinale")

# Step 2: Replace "Suface" with "Surface" inside text strings in the 'Model' column
df['Model'] = df['Model'].astype(str).str.replace('Suface', 'Surface', regex=False)


# Create 'Model Name' column
df['Model Name'] = df['Model'].apply(lambda x: str(x).split()[1] if isinstance(x, str) and len(str(x).split()) > 1 else '')
df.insert(df.columns.get_loc('Model') + 1, 'Model Name', df.pop('Model Name'))    

# Create 'Model Number' column
df['Model Number'] = df['Model'].apply(lambda x: str(x).split()[2] if isinstance(x, str) and len(str(x).split()) > 2 else '')
df.insert(df.columns.get_loc('Model Name') + 1, 'Model Number', df.pop('Model Number'))

# Create 'Form Factor' column
df['Form Factor'] = df['Description'].apply(lambda x: next((factor for factor in ['SFF', 'MFF', 'MT', 'Tower', 'AIO', 'NB'] if isinstance(x, str) and factor in x), 'NB'))
df.insert(df.columns.get_loc('Model Number') + 1, 'Form Factor', df.pop('Form Factor'))

# Create 'CPU' column
# df['CPU'] = df['Processor'].apply(lambda x: x[14:22] if isinstance(x, str) else '')
df['CPU'] = df['Processor'].apply(
lambda x: x[14:26] if isinstance(x, str) and 'Ultra' in x
else x[9:28] if isinstance(x, str) and 'Pentium' in x
else x[9:24] if isinstance(x, str) and 'Xeon' in x
else x[14:22] if isinstance(x, str) and 'Intel' in x
else x[7:20] if isinstance(x, str) and 'AMD' in x
else ''
)


# Create 'RAM' column
df['RAM'] = df['RAM'].apply(lambda x: x[:2] if isinstance(x, str) else x if isinstance(x, str) else x)

# Create 'ODD' column
df['ODD'] = df['Description'].apply(lambda x: 'DVDRW' if isinstance(x, str) and 'DVDRW' in x else ('DVD' if isinstance(x, str) and 'DVD' in x else ''))
df.insert(df.columns.get_loc('Form Factor') + 1, 'ODD', df.pop('ODD'))

# Create 'HDD' column
df['HDD'] = df['Description'].apply(lambda x: next((size for size in ['128GB', '256GB', '500GB', '512GB', '250GB', '320GB', '1000GB', '1024GB', '768GB'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('ODD') + 1, 'HDD', df.pop('HDD'))

# Create 'Screen Size' column
df['Screen_Size'] = df['Description'].apply(lambda x: next((size for size in ['14-in', '13.3-in','13.5-in', '12.3-in', '15.6-in','15,2','13,3','10-in','12.5-in','16-in','11.6-in','2N1','17.3','19.5-in', 'TAB', '24-in'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('HDD') + 1, 'Screen_Size', df.pop('Screen_Size'))

# Create 'Screen Type' column
df['Screen_Type'] = df['Description'].apply(lambda x: next((size for size in ['NoTCH', 'TCH'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('Screen_Size') + 1, 'Screen_Type', df.pop('Screen_Type'))

# Create 'CAMERA' column
df['Camera'] = df['Description'].apply(lambda x: next((size for size in ['NoCAM','CAM'] if isinstance(x, str) and size in x), ''))
df.insert(df.columns.get_loc('Screen_Type') + 1, 'Camera', df.pop('Camera'))

# Move 'CPU' and 'RAM' right after 'Model Number'
df.insert(df.columns.get_loc('Model Number') + 1, 'CPU', df.pop('CPU'))
df.insert(df.columns.get_loc('Model Number') + 2, 'RAM', df.pop('RAM'))

# Drop 'Processor' column
df.drop(columns=['Processor'], inplace=True)

# Sort the dataframe
df.sort_values(by=['Model Name', 'Model Number', 'Form Factor', 'CPU', 'RAM', 'HDD'], inplace=True)

# Create 'LISTS' directory if it doesn't exist
if not os.path.exists('LISTS'):
    os.makedirs('LISTS')

# Save the dataframe to Excel
current_time = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
new_file_path = os.path.join('LISTS', OUTPUT_FILENAME)
# new_file_path = os.path.join('LISTS', f'DELL_LIST_GRADE_A_{current_time}.xlsx')
df.to_excel(new_file_path, index=False)



# Load the workbook and adjust column widths
wb = load_workbook(new_file_path)
ws = wb.active
GREEN = '\033[92m'
RESET = '\033[0m'
# Add Prompt to calculate the 'Price' column
# Adjust column widths
for col in ws.columns:
    max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
    column_letter = col[0].column_letter
    ws.column_dimensions[column_letter].width = max_length + 2
# Rename columns
ws.cell(row=1, column=4, value='Model Number')  # Column D
ws.cell(row=1, column=14, value='Cost')  # Column M
# Add 'Price' column and calculate (column K / 0.75) + 10, formatted as currency
col_idx_k = 14
ws.insert_cols(col_idx_k + 1)
# assigning correct header to price or bid column
if ans1 == 'yes':
    ws.cell(row=1, column=col_idx_k + 1, value='Price')
else:
        ws.cell(row=1, column=col_idx_k + 1, value='Bid Price')
if ans1 == 'yes':
    # Placeholder logic for calculating the Price (you can replace this with actual logic)
    for row in range(2, ws.max_row + 1):
        cost = ws.cell(row=row, column=col_idx_k).value
        if isinstance(cost, (int, float)):
            price = round(((cost / 0.75) + 10)/5,0)*5
            price_cell = ws.cell(row=row, column=col_idx_k + 1, value=price)
            price_cell.number_format = '$#,##0.00'
else:
    # Leave the 'Price' column blank
    for row in range(2, ws.max_row + 1):
        price = ''
# MAKE EVERY LINE IN OS COLUMN = TO WINDOWS 10 PROFESSIONAL
col_idx_k = 16
# ws.insert_cols(col_idx_k + 1)
# ws.cell(row=1, column=col_idx_k + 1, value='Price')
for row in range(2, ws.max_row + 1):
    OS = ws.cell(row=row, column=col_idx_k).value
    if OS == 'No Operating System':
        OSN = 'Windows 10 Professional'
        ws.cell(row=row, column=col_idx_k).value=OSN
# MAKE EVERY LINE IN OPTIPLEX SHEET SCREEN SIZE = 'NO MONITOR'
col_idx_k = 10
# ws.insert_cols(col_idx_k + 1)
# ws.cell(row=1, column=col_idx_k + 1, value='Price')
for row in range(2, ws.max_row + 1):
    SCRN_S = ws.cell(row=row, column=col_idx_k).value
    if SCRN_S is None or SCRN_S == "":
        SCRN = 'Desktop Only'
        ws.cell(row=row, column=col_idx_k).value=SCRN

# MAKE EVERY LINE IN OPTIPLEX SHEET SCREEN TYPE = 'NO MONITOR'
col_idx_k = 11
# ws.insert_cols(col_idx_k + 1)
# ws.cell(row=1, column=col_idx_k + 1, value='Price')
for row in range(2, ws.max_row + 1):
    SCRN_T = ws.cell(row=row, column=col_idx_k).value
    if SCRN_T is None or SCRN_T == "":
        SCRN = 'Not Specified'
        ws.cell(row=row, column=col_idx_k).value=SCRN

# Delete 'Cost' and 'Product ID' 
cost_col_idx = 14
product_id_col_idx = 1  # Assuming 'Product ID' is the first column
# ws.delete_cols(cost_col_idx)
# ws.delete_cols(product_id_col_idx)
if ans2 == 'yes':
    # Assuming 'df' is the DataFrame being manipulated
    ws.delete_cols(cost_col_idx)
    ws.delete_cols(product_id_col_idx)
    print("'Cost' and 'Product ID' columns have been deleted.")
else:
    print("Columns 'Cost' and 'Product ID' will be retained.")
# Make J1 bold and center
cell_j1 = ws['J1']
cell_j1.font = Font(bold=True)
cell_j1.alignment = Alignment(horizontal='center')
# Save the intermediate workbook
wb.save(new_file_path)
# Re-load workbook to create tabs for each model name
# Load the Excel workbook
df = pd.read_excel(new_file_path, sheet_name='Sheet1')
# wb = load_workbook(new_file_path)
header = list(df.columns)
for model in df['Model Name'].unique():
    model_df = df[df['Model Name'] == model]
    model_ws = wb.create_sheet(title=model)
    # Write the header
    for col_idx, col_name in enumerate(header, 1):
        cell = model_ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    # Write the data
    for r_idx, row in enumerate(model_df.values, 2):
        for c_idx, value in enumerate(row, 1):
            model_ws.cell(row=r_idx, column=c_idx, value=value)
    # Adjust column widths for the new sheet
    for col in model_ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        if max_length > 0:
            column_letter = col[0].column_letter
            model_ws.column_dimensions[column_letter].width = max_length + 5
    # Format PRICE COLUMN as currency
    if ans2 == 'yes':
        for row in range(2, model_ws.max_row + 1):
            model_ws.cell(row=row, column=13).number_format = '$#,##0.00'
            # Set column width for column 13 (column 'M') to 10
            model_ws.column_dimensions[get_column_letter(13)].width = 10


    else:
        for row in range(2, model_ws.max_row + 1):
            model_ws.cell(row=row, column=14).number_format = '$#,##0.00'
            model_ws.column_dimensions[get_column_letter(14)].width = 10  # Column 'N'
        for row in range(2, model_ws.max_row + 1):    
            model_ws.cell(row=row, column=15).number_format = '$#,##0.00'
            model_ws.column_dimensions[get_column_letter(15)].width = 10  # Column 'O'

            
    # Delete column A
    # model_ws.delete_cols(1)
    # Autofit all columns
    for col in model_ws.columns:
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        column_letter = col[0].column_letter
        if max_length > 0:
            model_ws.column_dimensions[column_letter].width = max_length + 5
if ans2 == 'yes':
    # Set column J-I-K width
    model_ws.column_dimensions['L'].width = 15
    model_ws.column_dimensions['N'].width = 120
    model_ws.column_dimensions['M'].width = 30
else:
    # Set column J-I-K width
    model_ws.column_dimensions['L'].width = 15
    model_ws.column_dimensions['N'].width = 30
    model_ws.column_dimensions['M'].width = 15
    model_ws.column_dimensions['O'].width = 120
    # APPLY FILTERS
    model_ws.auto_filter.ref = ws.dimensions  # Applies filter to the range of dataws.auto_filter.ref = ws.dimensions  # Applies filter to the range of data
# Delete the original worksheet
wb.remove(wb['Sheet1'])
# Delete "Monitor" tab
for sheet_name in wb.sheetnames:
    if sheet_name in ["Monitor"]:
        wb.remove(wb[sheet_name])

    
# Save the final workbook
wb.save(new_file_path)




#********** merging all worksheets into the menu workbook ***********
# Define file paths
final_file_path = new_file_path  # Recently created workbook
menu_file_path = 'DELL_LIST_MENU.xlsx'  # Workbook with the "Menu" sheet
# --- FUNCTIONALITY TO MERGE BOTH WORKBOOKS ---
# Load the final workbook 'DFS_LIST_PROCESSED.xlsx' (the one with new sheets)
final_wb = load_workbook(final_file_path)
# Load the 'DELL_LIST_MENU.xlsx' workbook (which contains the Menu sheet)
menu_wb = load_workbook(menu_file_path)
# Copy each sheet from 'DFS_LIST_PROCESSED.xlsx' into 'DELL_LIST_MENU.xlsx'
for sheet_name in final_wb.sheetnames:
    # Get the current sheet from the final workbook
    source_sheet = final_wb[sheet_name]
    # If the sheet already exists in the menu workbook, remove it to avoid duplication
    if sheet_name in menu_wb.sheetnames:
        del menu_wb[sheet_name]
    # Create a new sheet in 'DELL_LIST_MENU.xlsx' with the same name as in 'DFS_LIST_PROCESSED.xlsx'
    target_sheet = menu_wb.create_sheet(sheet_name)
    # Copy each cell from the source sheet to the new target sheet, preserving values and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value
             # Copy cell font, fill, alignment, number_format, protection, and border
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.alignment = copy.copy(cell.alignment)
                new_cell.number_format = cell.number_format  # This is a simple string, no need to copy
                new_cell.protection = copy.copy(cell.protection)
                new_cell.border = copy.copy(cell.border)
    # Copy dimensions (column width, row height)
    for col in source_sheet.column_dimensions:
        target_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]
    for row in source_sheet.row_dimensions:
        target_sheet.row_dimensions[row] = source_sheet.row_dimensions[row]
    # Copy merged cells
    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))
 # --- Add Autofilter to the new sheet ---
    # Assuming the first row contains headers and we want to apply the autofilter over all data
    max_column = target_sheet.max_column  # Get the last column number with data
    max_row = target_sheet.max_row  # Get the last row number with data
    target_sheet.auto_filter.ref = f"A1:{target_sheet.cell(row=1, column=max_column).coordinate}"
    # The range is from A1 to the last column in the first row (header row)
# ADDING HYPER LINKS TO EACH TAB IN THE MENU
# Select the 'Menu' worksheet
menu_ws = menu_wb['MENU']
# --- Add hyperlinks under each image ---
# Add hyperlink for the Latitude tab under the first image
#menu_ws['N8'] = 'Go to Latitude'
#menu_ws['N8'].hyperlink = '#Latitude!A1'  # Hyperlink to the "Latitude" tab
#menu_ws['N8'].font = Font(color="0000FF", underline="single", size=16)  # Style it like a hyperlink
# Add hyperlink for the OptiPlex tab under the second image
#menu_ws['N10'] = 'Go to OptiPlex'
#menu_ws['N10'].hyperlink = '#OptiPlex!A1'  # Hyperlink to the "OptiPlex" tab
#menu_ws['N10'].font = Font(color="0000FF", underline="single", size=16)
# Add hyperlink for the Precision tab under the third image
#menu_ws['N12'] = 'Go to Precision'
#menu_ws['N12'].hyperlink = '#Precision!A1'  # Hyperlink to the "Precision" tab
#menu_ws['N12'].font = Font(color="0000FF", underline="single", size=16)
# CREATING SUMMARY TABLES FOR EACH TAB IN MENU SHEET
# Set the starting row for summary tables
# CREATING SUMMARY TABLES
# Set the starting row and column for the summary tables on the MENU tab
start_row = 10
start_col = 2  # Start from column A
# Define a border style for the tables
thin_border = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))
# Define center alignment
center_alignment = Alignment(horizontal='center', vertical='center')
# Define background fills for headers
blue_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
orange_fill = PatternFill(start_color="CC5500", end_color="CC5500", fill_type="solid")
green_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
black_fill = PatternFill(start_color="000000", end_color="008000", fill_type="solid")
# Define a white font for the headers
white_font = Font(color="FFFFFF", bold=True)
# Function to write a summary table to the MENU tab and adjust column width
def write_summary_to_menu(ws, df, sheet_name, start_row, start_col):
    print(f"Writing summary for {sheet_name} starting at row {start_row}, column {start_col}")
    # ###Add a total row to the DataFrame, summing the "Qty" column (assuming it exists)
    # Add a total row with the correct number of columns
    if "Qty" in df.columns:
        total_units = df["Qty"].sum()
        # Fill with "Total" for the first column, blanks for intermediate, and total in the last column
    total_row = pd.DataFrame(
            [["Total"] + [""] * (len(df.columns) - 2) + [total_units]],
            columns=df.columns)
    df = pd.concat([df, total_row], ignore_index=True)

    # Write the sheet name as a header, centered above the table
    header_cell = ws.cell(row=start_row, column=start_col, value=sheet_name)
    header_cell.font = Font(bold=True, size=14, color="FFFFFF")  # Set font size to 14 and bold
    header_cell.alignment = center_alignment
    header_cell.fill = black_fill
    #header_cell.font = white_font
    # Merge cells to center the header over summary columns
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=27)
    ws.cell(6,2, value ="SUMMARY TABLES BY CATEGORY, PROCESSOR AND FORM FACTOR")
    ws.cell(6,2).font = Font(bold=True, size=20, color="FFFFFF")  # Set font size to 14 and bold
    ws.cell(6,2).fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    ws.cell(6,2).alignment = center_alignment

    #SCALE VIEW TO 60%
    ws.sheet_view.zoomScale = 60

    
    # Merge cells to center the header over the table columns
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(df.columns) - 1)
    current_row = start_row + 1
    # Write the dataframe to the MENU tab, starting at the given column
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
        for c_idx, value in enumerate(row, start_col):
            cell = ws.cell(row=current_row + r_idx, column=c_idx, value=value)
            # Apply bold font and custom background fills for header row
            if r_idx == 0:
                cell.font = white_font  # Set white font
                # Apply background colors based on the column
                if c_idx == start_col:
                    cell.fill = blue_fill  # First header
                elif c_idx == start_col + 1:
                    cell.fill = orange_fill  # Second header
                elif c_idx == start_col + 2:
                    cell.fill = green_fill  # Third header
                elif c_idx == start_col + 3:
                    cell.fill = blue_fill  # Fourth header
                elif c_idx == start_col + 4:
                    cell.fill = green_fill  # Fifth header    
                # Apply border and center alignment to all cells
            cell.border = thin_border
            cell.alignment = center_alignment
    # Adjust column widths for the table (setting width to 20)
    for col_idx in range(start_col, start_col + len(df.columns)):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    # Return the next available column after writing the table
    return start_col + len(df.columns) + 2  # Adding 2 columns of space between tables
# Process each worksheet except the 'MENU' worksheet
for sheet_name in wb.sheetnames:
    if sheet_name != 'MENU':
        ws = wb[sheet_name]
        print(f"Processing sheet: {sheet_name}")

        # Load the data into a pandas DataFrame
        data = ws.values
        columns = next(data)[0:]  # First row is the header
        df = pd.DataFrame(data, columns=columns)
        # Check if required columns exist for summarization
        if 'CPU' in df.columns and 'Form Factor' in df.columns and 'Qty' in df.columns and 'Screen_Size' in df.columns and 'Screen_Type' in df.columns:
            print(f"Creating summary for {sheet_name}")
            # Create summary table grouped by CPU and Form Factor, with total qty
            summary_df = df.groupby(['CPU', 'Form Factor', 'Screen_Size', 'Screen_Type']).agg({'Qty': 'sum'}).reset_index()
            # Write the summary table to the MENU tab
            start_col = write_summary_to_menu(menu_ws, summary_df, sheet_name, start_row, start_col)
        else:
            print(f"Required columns not found in {sheet_name}")





# Save the updated 'DELL_LIST_MENU.xlsx' file, which now contains all sheet

# Prepare the file in memory
excel_stream = BytesIO()
menu_wb.save(excel_stream)
excel_stream.seek(0)

# Upload to Dropbox (overwrite mode)
dropbox_path = f"/AUTODFSMAILER/LISTS/{OUTPUT_FILENAME}"  # assuming OUTPUT_FILENAME = 'DELL_LIST_GRADE_A_LATEST.xlsx'
dbx.files_upload(excel_stream.read(), dropbox_path, mode=WriteMode("overwrite"))


#dbx.files_upload(excel_stream.read(), OUTPUT_PATH, mode=WriteMode("overwrite"))
print(f"Saved processed file to Dropbox at: {OUTPUT_PATH}")
from io import BytesIO

# Save the same file locally so it can be uploaded to Google Drive
excel_stream.seek(0)
final_file_path = f"/tmp/{OUTPUT_FILENAME}"
with open(final_file_path, "wb") as f:
    f.write(excel_stream.getbuffer())

# Upload to Google Drive
google_drive_file_id = upload_to_drive_oauth(
    file_path=final_file_path,
    file_name=os.path.basename(final_file_path)
)


print(f"Saved file to Google Drive with ID: {google_drive_file_id}")

# Output the message with the total number of lines
total_lines = len(df)  # Total number of lines in the original dataframe
print(f"Done Processing... Total number of lines: {total_lines}")


# --- Email the file using Gmail (from the same final_file_path) ---
print("Sending final file with gmail to yosi@myy-tech.com")
import smtplib
import ssl
from email.message import EmailMessage
import mimetypes
from pathlib import Path

sender_email = os.getenv("GMAIL_SENDER_EMAIL")
app_password = os.getenv("GMAIL_APP_PASSWORD")
recipient_email = os.getenv("RECIPIENT_EMAIL")


# Setup email
msg = EmailMessage()
msg["Subject"] = "Final DELL List"
msg["From"] = sender_email
msg["To"] = recipient_email
msg.set_content("Please find the final DELL list attached.")

# Attach the local file
filepath = final_file_path  # should be '/tmp/DELL_LIST_GRADE_A_LATEST.xlsx'
mime_type, _ = mimetypes.guess_type(filepath)
maintype, subtype = mime_type.split("/")

with open(filepath, "rb") as f:
    msg.add_attachment(
        f.read(),
        maintype=maintype,
        subtype=subtype,
        filename=Path(filepath).name
    )

# Send email
context = ssl.create_default_context()
with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context) as server:
    server.login(sender_email, app_password)
    server.send_message(msg)

print("Final file emailed successfully!")




# Open the final Excel file
#try:
 #   subprocess.run(["open", new_file_path])  # Use 'open' for macOS
 #   print(f"Opened {new_file_path} successfully.")
# except Exception as e:
    

# Create the GUI
# def run_with_answers(ans1, ans2):
#   root.destroy()
#   main(ans1, ans2)
