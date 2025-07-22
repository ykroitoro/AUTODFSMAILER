
import os
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from openpyxl.drawing.image import Image
import copy  # Import the copy module for deep copying

# Define file path
file_path = 'DFS_LIST.xlsx'

# Load the Excel workbook
df = pd.read_excel(file_path, sheet_name='Worksheet')

# Create 'Model Name' column
df['Model Name'] = df['Model'].apply(lambda x: x.split()[1] if len(x.split()) > 1 else '')
df.insert(df.columns.get_loc('Model') + 1, 'Model Name', df.pop('Model Name'))

# Create 'Model Number' column
df['Model Number'] = df['Model'].apply(lambda x: x.split()[2] if len(x.split()) > 2 and x.split()[2].isdigit() else (x.split()[2] if len(x.split()) > 2 else ''))
df.insert(df.columns.get_loc('Model Name') + 1, 'Model Number', df.pop('Model Number'))

# Create 'Form Factor' column
df['Form Factor'] = df['Description'].apply(lambda x: next((factor for factor in ['SFF', 'MFF', 'MT', 'Tower', 'AIO', 'NB'] if isinstance(x, str) and factor in x), 'NB'))
df.insert(df.columns.get_loc('Model Number') + 1, 'Form Factor', df.pop('Form Factor'))

# Create 'CPU' column
df['CPU'] = df['Processor'].apply(lambda x: x[14:22] if isinstance(x, str) else '')
df['RAM'] = df['RAM'].apply(lambda x: x[:2] if isinstance(x, str) else x)

# Create 'ODD' column
df['ODD'] = df['Description'].apply(lambda x: 'DVDRW' if isinstance(x, str) and 'DVDRW' in x else ('DVD' if isinstance(x, str) and 'DVD' in x else ''))

# Add Prompt to calculate the 'Price' column
user_choice = input("Do you want to calculate the Price column? (yes/no): ").strip().lower()

if user_choice == 'yes':
    # Placeholder logic for calculating the Price (you can replace this with actual logic)
    df['Price'] = df.apply(lambda row: 100, axis=1)  # Example calculation logic
    print("Price column has been calculated.")
else:
    # Leave the 'Price' column blank
    df['Price'] = ''
    print("Price column has been left blank.")

# Continue with your script logic for saving the Excel workbook...

# Copy each cell from the source sheet to the new target sheet, preserving values and styles
# ... rest of the existing code ...

