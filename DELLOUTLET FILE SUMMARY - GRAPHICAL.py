import tkinter as tk
from tkinter import filedialog
import pandas as pd
import os
import openpyxl
import xlrd
import xlsxwriter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime
#from UliPlot.XLSX import auto_adjust_xlsx_column_width
from typing import NoReturn

os.system('cls' if os.name == 'nt' else 'clear')

def run_script():
    os.system('cls' if os.name == 'nt' else 'clear')
    print("Script running...")



# Read the data from the DELLOUTLET.XLSX workbook
    df = pd.read_excel('DELLOUTLET.XLSX')

# Rename the 'OUTLET Sku' column to 'Outlet SKU'
#df = df.rename(columns={'OUTLET Sku': 'Outlet SKU'})

# Group the data by 'Outlet SKU'
    grouped = df.groupby('Outlet SKU')

# Summarize the data by adding the total number of related 'Service Tag' and 
# calculating the average of the 'Reseller Price' column
    result = grouped['Service Tag'].count().reset_index()
    result['Reseller Price '] = grouped['Reseller Price '].mean().reset_index()['Reseller Price ']

# Add all the columns from the original file to the summarized data
    result = pd.merge(result, df, on='Outlet SKU', how='left')

# Drop duplicates from the summarized data
    result = result.drop_duplicates(subset='Outlet SKU')

# Rename the 'Service Tag_x' column to 'Qty'
    result = result.rename(columns={'Outlet SKU': 'SKU#'})

# Rename the 'Service Tag_x' column to 'Qty'
    result = result.rename(columns={'Service Tag_x': 'Qty'})

# Rename the 'Reseller Price_x' column to 'Cost'
    result = result.rename(columns={'Reseller Price _x': 'Cost'})

# Rename the 'Product' column to 'MPN'
    result = result.rename(columns={'Product': 'MPN'})

# Rename the 'Processor' column to 'CPU'
    result = result.rename(columns={'Processor': 'CPU'})

# Rename the 'Graphics' column to 'VIDEO CARD'
    result = result.rename(columns={'Graphics': 'VIDEO CARD'})

# Rename the 'Media Bay' column to 'ODD'
    result = result.rename(columns={'Media Bay': 'ODD'})

# Rename the 'Mem Total' column to 'MEM'
    result = result.rename(columns={'Mem Total': 'MEM'})

# Rename the 'Default Warranty' column to 'WARRANTY'
    result = result.rename(columns={'Default Warranty': 'WARRANTY'})

# Rename the 'Display' column to 'DISPLAY'
    result = result.rename(columns={'Display': 'DISPLAY'})

# Drop the 'Service Tag_y' and 'Reseller Price_y' columns
    result = result.drop(columns=['Service Tag_y', 'Reseller Price _y','Condition','Prc. Qty','Networking','Chassis','Battery','Camera','Keyboard','Mouse / Touchpad','Color','Power','Feature 1','Feature 2','Feature 3','Feature 4','Feature 5','HARDWARE_UPGRADE','FAMILY_NAME','FAMILY_NAME','LOB','PRICE_NEW','FAMILY_SERIES'])

# Calculate the 'Price' column with the formula round(round(cost/.85,0)/5,0)*5
    result['Price'] = round(round(result['Cost']/.85,0)/5,0)*5

# Rename the 'Price' column to 'PRICE'
    result = result.rename(columns={'Price': 'PRICE'})


# Write the summarized data to a new Excel file
    result.to_excel('DELL_LIST_GPT.xlsx', index=False)

#OPEN THE NEW WORKBOOK AND REARRANGE THE COLUMNS

# Read the excel file
    df = pd.read_excel("DELL_LIST_GPT.xlsx")

# Drop the 'Cost' column
    df = df.drop(columns=["Cost"])

# Rearrange the columns in the specified order
    df = df[['Segment', 'SKU#', 'MPN', 'Qty', 'PRICE', 'CPU', 'MEM', 'HDD', 'ODD', 'OS', 'DISPLAY', 'VIDEO CARD', 'WARRANTY']]


# Remove the word "Outlet" from the "MPN" column
    df["MPN"] = df["MPN"].str.replace("Outlet", "")

# Write the dataframe to an excel file
    df.to_excel("DELL_LIST.xlsx", index=False)




# PART 2 - FORMATING DATA INTO A NEW EXCEL WORKBOOK WITH MULTIPLE TABS, FOR EACH SEGMENT.


    #import pandas as PD

#CREATING MASTER EXCEL WITH MULTIPLE TAB FOR EACH CATEGORY


   # Open the "DELL_LIST.xlsx" file using openpyxl
    wb = openpyxl.load_workbook("DELL_LIST.xlsx")

    # Get the first sheet in the workbook
    ws = wb[wb.sheetnames[0]]

    # Get the list of unique Segment values from the sheet
    segments = list(set(row[0].value for row in ws.iter_rows(min_row=2, max_col=1, values_only=True)))

    # Create a new workbook for the sorted data
    new_wb = openpyxl.Workbook()

    # Loop through the unique Segment values
    for segment in segments:
        # Create a new sheet for the current Segment
        new_ws = new_wb.create_sheet(title=segment)

        # Write the header row to the new sheet
        new_ws.append(["MPN", "Title", "Segment"])

        # Loop through the rows in the original sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check if the current row matches the current Segment
            if row[0].value == segment:
                # Write the current row to the new sheet, sorted by MPN
                new_ws.append([row[2], row[1], row[0]])

    # Save the new workbook
    new_wb.save("DELLTODAY.xlsx")




# FORMATING THE EXCEL TO CREATE TABLES IN EACH WORKSHEET WITH FORMATED LOOK AND FILTERS

    from openpyxl import load_workbook
    wb = load_workbook("DELLTODAY.xlsx")

    sheets = wb.sheetnames
    i=1
    for s_name in sheets:
        #print(s_name)
        # strip blank spaces from table name
        tablename = s_name.replace(' ','')
        sheet=wb[s_name]
    
    
        # Get the dimensions of the dataframe.
        (max_row, max_col) = df.shape
        #location of last cell in the current sheet
        lastcell = xlsxwriter.utility.xl_col_to_name(max_col)+str(max_row)
        # Create a list of column headers, to use in add_table().
        column_settings = [{'header': column} for column in df.columns]
        # Add the Excel table structure. Pandas will add the data.
        table = Table(displayName = tablename, ref="A1:" + get_column_letter(sheet.max_column) + str(sheet.max_row))
        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        sheet.add_table(table)

        #CHANGE PRICE ROW HEADER
        sheet['E1']='PRICE'
        #FORMAT COLUMN E - PRICE COLUMN AS USD
        s_max = sheet.max_row
        for row in range(2,s_max+1):
            s_row = ("E"+ str(row))
            sheet[("E" + str(row))].number_format = '$#,##0_-'

   
    # Make the columns wider for clarity.
    #Autofit column in active sheet
    
    # Auto-adjust columns' width
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 25
        sheet.column_dimensions['C'].width = 35
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 10
        sheet.column_dimensions['F'].width = 75
        sheet.column_dimensions['G'].width = 20
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 30
        sheet.column_dimensions['J'].width = 20
        sheet.column_dimensions['K'].width = 75
        sheet.column_dimensions['L'].width = 100

        donesheet = (s_name+'- Processed')
        print(donesheet)


# INSERT INFORMATION IMAGE
# INSERT INFORMATION SHEET IMAGE
    sheet2 = wb.create_sheet(title="INFO")

#sheet=wb['INFORMATION']


    sheet = wb.copy_worksheet(wb["INFO"])
    sheet.title = "INFORMATION"
    wb.move_sheet("INFORMATION", -(len(wb.sheetnames)-1))
    img = openpyxl.drawing.image.Image('INFORMATION.jpg')
    img.anchor = 'A1'
    sheet.add_image(img)

    del wb['INFO']


    wb.save('DELL_FACTORY_CERTIFIED_'+date_time_str+'.xlsx')

    print('')
    print('PROCESS COMPLETED')


root = tk.Tk()
#root.config(bg='gray')
root.geometry("400x300") # Set the size of the window to 400x300 pixels

root.title("DELL OUTLET PYTHON PROCESSING SCRIPT")


def cancel_script():
    # Your code for cancel button here
    print("Script cancelled")
    root.destroy()
    os.system('clear')


run_button = tk.Button(root, text="Run script", command=run_script)
cancel_button = tk.Button(root, text="Cancel", command=cancel_script)

run_button.pack()
cancel_button.pack()




root.mainloop()



